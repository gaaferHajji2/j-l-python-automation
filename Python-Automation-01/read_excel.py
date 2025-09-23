from openpyxl import load_workbook

workbook = load_workbook('jloka_test_01.xlsx');

print("Sheet Names Are: ", str(workbook.sheetnames));

sheet = workbook.active

print("The value of A1 cell: ", sheet['A1'].value);

print("The cell value is: ", sheet.cell(row=1, column=1).value)

print("The values using column: ", sheet["A:B"])

print("The values of row 1: ", sheet[1])

for value in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print("The values of rows: ", str(value))

for value in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print("The values of cols: ", str(value))

workbook.close();